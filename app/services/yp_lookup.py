"""YP (Young Professional) lookup service.

Loads a YP allocation file once at startup and provides an in-memory,
case-insensitive lookup by MDO / org-channel name.

File is configured via YP_ALLOCATION_FILE env var — never committed to git.

Supported formats (detected by extension):
  .csv   — preferred; stdlib only, no extra dependencies
           Columns: centre_state, mdo, name, email, mobile, cc_email  (header row)
  .xlsx  — legacy Excel; requires openpyxl
           Column order: A=centre_state  B=mdo  C=name  D=email  E=mobile  F=cc_email

If YP_ALLOCATION_FILE is not set or the file is missing, the service starts
with an empty index and logs a warning — no crash, flow degrades gracefully.

To convert the existing Excel to CSV (one-time, run locally):
    python -c "
    import csv, openpyxl
    wb = openpyxl.load_workbook('data/Allocation_28.10.2025.xlsx', read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    with open('yp_allocation.csv', 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['centre_state','mdo','name','email','mobile','cc_email'])
        for r in rows[1:]:
            if r[1]: w.writerow([r[0] or '', r[1] or '', r[2] or '', r[3] or '', r[4] or '', r[5] or ''])
    "
Then deploy yp_allocation.csv to the server and set YP_ALLOCATION_FILE=/path/to/yp_allocation.csv
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any

from app.config import settings

log = logging.getLogger(__name__)


class YPLookupService:
    """In-memory lookup service backed by the YP allocation file.

    After construction the index is ready — no I/O at query time.
    Starts empty (with a warning) if YP_ALLOCATION_FILE is not configured.
    """

    def __init__(self) -> None:
        self._index: dict[str, dict[str, Any]] = {}
        self._load()

    def _load(self) -> None:
        path_str = settings.yp_allocation_file.strip()
        if not path_str:
            log.warning("YP_ALLOCATION_FILE not set — YP lookup will return nothing")
            return

        path = Path(path_str)
        if not path.exists():
            log.warning("YP allocation file not found: %s", path)
            return

        ext = path.suffix.lower()
        if ext == ".csv":
            self._load_csv(path)
        elif ext in (".xlsx", ".xls"):
            self._load_xlsx(path)
        else:
            log.error("Unsupported YP file format %r — use .csv or .xlsx", ext)

    def _load_csv(self, path: Path) -> None:
        try:
            with path.open(newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                count = 0
                for row in reader:
                    mdo_raw = row.get("mdo", "").strip()
                    if not mdo_raw:
                        continue
                    self._index[mdo_raw.lower()] = {
                        "centre_state": row.get("centre_state", "").strip(),
                        "mdo":          mdo_raw,
                        "name":         row.get("name", "").strip(),
                        "email":        row.get("email", "").strip(),
                        "mobile":       row.get("mobile", "").strip(),
                        "cc_email":     row.get("cc_email", "").strip(),
                    }
                    count += 1
            log.info("YP lookup loaded %d entries from %s", count, path.name)
        except Exception as exc:
            log.error("YP CSV load failed: %s", exc)

    def _load_xlsx(self, path: Path) -> None:
        try:
            import openpyxl
        except ImportError:
            log.error("openpyxl not installed — cannot read .xlsx; convert to .csv instead")
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            count = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # skip header
                mdo_raw = row[1]
                if not mdo_raw:
                    continue
                self._index[str(mdo_raw).strip().lower()] = {
                    "centre_state": str(row[0]).strip() if row[0] else "",
                    "mdo":          str(mdo_raw).strip(),
                    "name":         str(row[2]).strip() if row[2] else "",
                    "email":        str(row[3]).strip() if row[3] else "",
                    "mobile":       str(row[4]).strip() if row[4] else "",
                    "cc_email":     str(row[5]).strip() if row[5] else "",
                }
                count += 1
            wb.close()
            log.info("YP lookup loaded %d entries from %s", count, path.name)
        except Exception as exc:
            log.error("YP XLSX load failed: %s", exc)

    def lookup(self, org_channel: str | None) -> dict[str, Any] | None:
        """Return the YP entry for *org_channel*, or ``None`` if not found.

        Match is case-insensitive and strips whitespace.
        Returns dict with keys: centre_state, mdo, name, email, mobile, cc_email.
        """
        if not org_channel:
            return None
        return self._index.get(str(org_channel).strip().lower())

    def __len__(self) -> int:
        return len(self._index)
